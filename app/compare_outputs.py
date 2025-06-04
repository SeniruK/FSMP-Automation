# Functions for comparing FSMP years

def run_all_calculations(FSMPyear, table_names):
    from .database import get_table_data
    from .calculateFSMP import calculateIATModifiedTotalLife, calculateCPDTLife, calculateCPFollowonLife, calculateInspectionIntervals, calculateHoursAtNextInspection, calculateHoursToNextInspection
    import pandas as pd

    db_dataframes = {}
    for name in table_names:
        data = get_table_data(name, FSMPyear)
        df = pd.DataFrame(data)
        db_dataframes[name] = df
    
    fsmp_dataframes = {}
    fsmp_dataframes["iat_modified_total_life"] = calculateIATModifiedTotalLife(FSMPyear, db_dataframes, fsmp_dataframes)
    fsmp_dataframes["cp_dt_life"] = calculateCPDTLife(FSMPyear, db_dataframes, fsmp_dataframes)
    fsmp_dataframes["cp_follow_on_life"] = calculateCPFollowonLife(FSMPyear, db_dataframes, fsmp_dataframes)
    fsmp_dataframes["inspection_intervals"] = calculateInspectionIntervals(FSMPyear, db_dataframes, fsmp_dataframes)
    fsmp_dataframes["hours_at_next_inspection"] = calculateHoursAtNextInspection(FSMPyear, db_dataframes, fsmp_dataframes)
    fsmp_dataframes["hours_to_next_inspection"] = calculateHoursToNextInspection(FSMPyear, db_dataframes, fsmp_dataframes)

    return fsmp_dataframes

def compareOutputs(df_current, df_previous, filename):
    import pandas as pd
    import numpy as np
    import openpyxl

    previous_acsn_list = df_previous["ACSN"].tolist()
    previous_columns = df_previous.columns

    df_current_filtered = df_current[previous_columns]
    df_current_filtered = df_current_filtered[df_current_filtered["ACSN"].isin(df_previous["ACSN"])]

    df_previous2 = df_previous.set_index("ACSN")
    df_previous2 = df_previous2.astype(np.float64)
    df_current_filtered2 = df_current_filtered.set_index("ACSN")
    df_current_filtered2 = df_current_filtered2.astype(np.float64)
    df_diff = df_previous2 - df_current_filtered2

    highlight_list = []
    for col in df_diff.columns:
        if (round(df_diff[col]) != 0).any():
            highlight_list.append(col)
    # print(highlight_list)

    # excel_file = f"Compare/{FSMPyear}/COMPARE {comparison_sheet}.xlsx"
    with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
        df_python_filtered.to_excel(writer, sheet_name="Python", index=False)
        df_original.to_excel(writer, sheet_name="Original", index=False)

    workbook = openpyxl.load_workbook(excel_file)
    worksheet = workbook.create_sheet(title="COMPARE")

    for index, value in enumerate(df_original['ACSN'], start=1):
        worksheet.cell(row=index+1, column=1, value=value)

    # Write headers
    for col_num, column_title in enumerate(df_original.columns, start=1):
        worksheet.cell(row=1, column=col_num, value=column_title)
    
    for cell in worksheet[1:1]:
        if cell.value in highlight_list:
            cell.fill = openpyxl.styles.PatternFill(start_color="fcc200", end_color="fcc200", fill_type="solid")
        cell.font = openpyxl.styles.Font(bold=True)
        cell.alignment = openpyxl.styles.Alignment(horizontal="center")
        cell.border = openpyxl.styles.Border(top=openpyxl.styles.Side(border_style="thin"), left=openpyxl.styles.Side(border_style="thin"), right=openpyxl.styles.Side(border_style="thin"), bottom=openpyxl.styles.Side(border_style="thin"))

    # Write formulas to the 'COMPARE' sheet
    for row in range(2, len(df_original) + 2):  # Start from row 2 to account for header
        for col in range(2, len(df_original.columns) + 1):  # Start from column 1 (A)
            cell_formula = f'=IF(OR(Original!{openpyxl.utils.get_column_letter(col)}{row}="", Python!{openpyxl.utils.get_column_letter(col)}{row}=""),"",Original!{openpyxl.utils.get_column_letter(col)}{row}-Python!{openpyxl.utils.get_column_letter(col)}{row})'
            worksheet[f"{openpyxl.utils.get_column_letter(col)}{row}"].value = cell_formula
            # worksheet.cell(row=row, column=col, value=cell_formula)
            worksheet[f"{openpyxl.utils.get_column_letter(col)}{row}"].number_format = "0"

    
    workbook.save(filename)